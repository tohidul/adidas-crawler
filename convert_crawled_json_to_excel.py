import openpyxl
import json
from openpyxl.styles import Font

def get_crawled_data_dict(crawled_data_file_path:str):
    with open(crawled_data_file_path,  encoding='UTF8') as json_file:
        crawled_data = json.load(json_file)

    return crawled_data


def get_chart_matrix(size_chart_list):
    size_chart_matrix = []
    temp_size_chart_list = []
    found_all_chart = 0
    total_number_of_size = -1
    for elem in size_chart_list:
        if 'cm' in elem and found_all_chart == 0:
            found_all_chart = 1
            total_number_of_size = len(temp_size_chart_list)
            size_chart_matrix.append(temp_size_chart_list)
            temp_size_chart_list = []
        if found_all_chart and total_number_of_size == len(temp_size_chart_list):
            size_chart_matrix.append(temp_size_chart_list)
            temp_size_chart_list = []
        temp_size_chart_list.append(elem)
    else:
        if temp_size_chart_list:
            size_chart_matrix.append(temp_size_chart_list)
    return size_chart_matrix

def write_to_excel(crawled_data:dict, excel_path:str):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "adidas_data"

    columns = ["product_url", "product_name", "category", "price", "available_size", "breadcrumb", "title_of_description","general_description_of_the_product", "general_description_itemized","rating","number_of_reviews","recommend_rate", "size"]
    for i, column in enumerate(columns):
        cell = sheet.cell(row=1, column=i+1)
        cell.value = column
        cell.font = Font(bold=True)
        if column=='size':
            sheet.merge_cells(start_row=1, start_column=i+1, end_row=1, end_column=i+20)

    row_num = 2
    for product in crawled_data:
        for attribute in product.keys():
            if attribute not in ["size_header", "size_chart"]:
                attribute_index = columns.index(attribute)
                cell = sheet.cell(row=row_num, column=attribute_index+1)
                cell.value = str(product[attribute])

        size_header_list = product["size_header"]
        size_chart_list = product["size_chart"]
        size_chart_matrix = get_chart_matrix(size_chart_list)

        for i, size_chart in enumerate(size_chart_matrix):
            for r in range(len(size_chart)):
                cell = sheet.cell(row=row_num+i, column=columns.index("size")+r+2)
                cell.value=size_chart_matrix[i][r]

        for i,size_header in enumerate(size_header_list):
            cell = sheet.cell(row=row_num + i+1, column=columns.index("size") + 1)
            cell.value = size_header_list[i]
        row_num+=len(size_header_list)+1




    wb.save(excel_path)


if __name__=='__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Convert json to excel')
    parser.add_argument('json_path',
                        type=str,
                        help='crawled json path')
    parser.add_argument('excel_path',
                        type=str,
                        help='excel output')
    args = parser.parse_args()
    crawled_dict = get_crawled_data_dict(args.json_path)
    write_to_excel(crawled_data=crawled_dict,excel_path=args.excel_path)




